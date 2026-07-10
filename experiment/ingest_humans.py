#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ingest_humans.py — 鲁棒读取乱格式的人类问卷 xlsx,判分,输出统一长表 + 解析 QA。
用法:  python3 ingest_humans.py
读:   human_samples/human_data_set1_zh_ans/*.xlsx
出:   results/human_long.csv   (schema 与模型 master 对齐,model=human,带 participant_id)
      并打印每个文件的解析情况(读到多少题、答案/置信度缺多少),标出需人工看的文件。
"""
import glob, os, csv, sys, re
from openpyxl import load_workbook
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import run_experiment as R

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "human_samples", "human_data_set1_zh_ans")
BANK = os.path.join(HERE, "..", "题库_ItemBank_v2_400.xlsx")

# 题库:item_id -> {gold,type,subtype,difficulty}
ITEMS = {it["id"]: it for it in R.load_items(BANK, "ItemBank 题库")}

TRUE = {"对", "正确", "true", "t", "√", "yes", "y", "是", "1", "✓", "√"}
FALSE = {"错", "错误", "false", "f", "×", "x", "no", "n", "否", "0"}

def norm_answer(v):
    if v is None: return None
    s = str(v).strip().lower()
    if s in TRUE: return "True"
    if s in FALSE: return "False"
    if s.startswith(("对","正","t","√","y","是")): return "True"
    if s.startswith(("错","f","×","x","n","否")): return "False"
    return None

def norm_conf(v):
    if v is None: return None
    s = str(v).strip()
    m = re.search(r"[1-5]", s)
    return int(m.group(0)) if m else None

def find_header(ws):
    """找表头行 & 列索引(题号/判断/置信度)。返回 (header_row, col_item, col_ans, col_conf) 或 None。"""
    for r in range(1, min(ws.max_row, 15) + 1):
        texts = {c: str(ws.cell(row=r, column=c).value or "") for c in range(1, ws.max_column + 1)}
        joined = "".join(texts.values())
        if "题号" in joined and ("确信" in joined or "置信" in joined):
            ci = ca = cc = None
            for c, t in texts.items():
                if "题号" in t: ci = c
                if "判断" in t or "对/错" in t or "答案" in t: ca = c
                if "确信" in t or "置信" in t: cc = c
            if ci and cc:
                return r, ci, (ca or ci + 2), cc
    return None

rows_out = []
qa = []
for f in sorted(glob.glob(os.path.join(SRC, "*.xlsx"))):
    fn = os.path.basename(f)
    pid = fn.rsplit(".", 1)[0]           # participant_id = 文件名(去扩展名)
    try:
        ws = load_workbook(f, data_only=True).active
    except Exception as e:
        qa.append((fn, 0, 0, 0, f"打开失败:{e}")); continue
    hdr = find_header(ws)
    if not hdr:
        qa.append((fn, 0, 0, 0, "找不到表头(题号/确信度)")); continue
    hr, ci, ca, cc = hdr
    n_item = n_ans = n_conf = n_scored = 0
    file_rows = []
    for r in range(hr + 1, ws.max_row + 1):
        iid = ws.cell(row=r, column=ci).value
        if iid is None: continue
        iid = str(iid).strip()
        if iid not in ITEMS: continue        # 不是有效题号则跳过
        n_item += 1
        ans = norm_answer(ws.cell(row=r, column=ca).value)
        conf = norm_conf(ws.cell(row=r, column=cc).value)
        if ans is not None: n_ans += 1
        if conf is not None: n_conf += 1
        it = ITEMS[iid]
        correct = "" if ans is None else str(ans == it["gold"])
        ok = ans is not None and conf is not None
        if ok: n_scored += 1
        file_rows.append({
            "model": "human", "model_id": fn, "served_model": "", "participant_id": pid,
            "language": "zh", "item_id": iid, "type": it["type"], "subtype": it["subtype"],
            "difficulty": it["difficulty"], "gold": it["gold"], "sample_idx": 0,
            "parsed_answer": ans or "", "correct": correct,
            "confidence": conf if conf is not None else "",
            "parse_ok": ok, "status": "ok" if ok else "fail_parse",
        })
    if n_ans == 0:                        # 完全空白的问卷,跳过不入库
        qa.append((fn, n_item, n_ans, n_conf, "空白未填 → 跳过")); continue
    rows_out.extend(file_rows)
    flag = "" if (n_item >= 30 and n_ans >= n_item * 0.9 and n_conf >= n_item * 0.9) else "  ⚠️需看"
    qa.append((fn, n_item, n_ans, n_conf, f"有效{n_scored}{flag}"))

out = os.path.join(HERE, "results", "human_long.csv")
cols = ["model","model_id","served_model","participant_id","language","item_id","type","subtype",
        "difficulty","gold","sample_idx","parsed_answer","correct","confidence","parse_ok","status"]
with open(out, "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=cols); w.writeheader(); w.writerows(rows_out)

print(f"读入 {len(qa)} 个文件,共写 {len(rows_out)} 行 → {out}\n")
print(f"{'文件':<40}{'题':>4}{'答':>5}{'信度':>5}  备注")
for fn, ni, na, nc, note in qa:
    print(f"{fn:<40}{ni:>4}{na:>5}{nc:>5}  {note}")
n_part = len(set(r["participant_id"] for r in rows_out))
print(f"\n被试(文件)数:{n_part} | 有效行:{sum(1 for r in rows_out if r['parse_ok'])}/{len(rows_out)}")
