#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_experiment.py — 元认知校准研究:模型批量测评运行器
Metacognitive-calibration study: batch runner for local + API models.

设计要点:
- 统一 OpenAI 兼容客户端:LM Studio(本地)与各家 API 共用同一套代码,只在 models.yaml 里改 base_url/api_key。
- 无上下文记忆:每题一次独立调用(仅 system + 当题 user),题与题之间互不影响。
- 每题多次采样(temperature>0),也可在配置里改成 temperature=0、n_samples=1 做确定性单次。
- 增量写 CSV + 断点续跑:重复运行会跳过已完成的 (model, language, item_id, sample_idx)。
- 稳健解析:优先解析 JSON,失败则回退关键词解析,标记 parse_ok。
- 可选 logprobs:辅助分析(仅支持的模型),存原始 logprobs,不进主对比。

用法示例:
  pip install -r requirements.txt
  # 先小规模自测(只跑前 5 题、单个模型):
  python run_experiment.py --limit 5 --models local-gemma
  # 正式全量:
  python run_experiment.py
  # 只跑中文、指定两个模型:
  python run_experiment.py --languages zh --models gpt-4o-mini,deepseek-chat
  # 解析器自测(不调用任何模型):
  python run_experiment.py --selftest
"""
import argparse, csv, json, os, re, sys, time, datetime, threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

try:
    import yaml
except ImportError:
    sys.exit("缺少依赖:pip install pyyaml")
try:
    from openai import OpenAI
except ImportError:
    sys.exit("缺少依赖:pip install openai")
try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("缺少依赖:pip install openpyxl")

HERE = Path(__file__).resolve().parent

def load_dotenv(path):
    """轻量 .env 加载器(无需第三方库):把本目录 .env 里的 KEY=VALUE 注入环境变量。
    已存在的真实环境变量优先,不被覆盖。"""
    if not os.path.exists(path):
        return
    for line in open(path, encoding="utf-8"):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and v and k not in os.environ:
            os.environ[k] = v

# ---------- prompt 模板(与 standard_prompt.md 保持一致) ----------
SYS = {
 "en": "You are answering a series of independent True/False questions. For each statement, first decide whether it is True or False, and then report how confident you are in that answer. Rely only on your own knowledge. You will not be told whether you are correct, and each question is independent of the others.",
 "zh": "你正在回答一系列相互独立的判断题。对每个陈述,请先判断它是“对(True)”还是“错(False)”,然后报告你对这个答案的确信程度。仅依据你自己的知识作答。你不会被告知答案是否正确,每道题之间相互独立。",
}
USER = {
 "en": ("Statement: {s}\n\n"
        "First decide whether the statement is True or False.\n"
        "Then rate your confidence in that answer using exactly one of these five levels:\n"
        "1 = pure guess\n2 = not very sure\n3 = half-and-half\n4 = fairly sure\n5 = very sure\n\n"
        "Respond with ONLY a raw JSON object — no markdown, no code fences, no extra text — "
        "in exactly this format:\n"
        '{{"answer": "True or False", "confidence": 1}}'),
 "zh": ("陈述:{s}\n\n"
        "请先判断该陈述是“对(True)”还是“错(False)”。\n"
        "然后用以下五档中的恰好一档,报告你的确信程度:\n"
        "1 = 完全靠猜 (pure guess)\n2 = 不太确定 (not very sure)\n3 = 一半一半 (half-and-half)\n4 = 比较确定 (fairly sure)\n5 = 非常确定 (very sure)\n\n"
        "只输出纯 JSON 对象 —— 不要 markdown、不要代码围栏、不要任何多余文字 —— 严格采用如下格式:\n"
        '{{"answer": "True 或 False", "confidence": 1}}'),
}
LANG_COL = {"en": "en", "zh": "zh"}  # 题干取哪一列

CSV_FIELDS = ["timestamp","model","model_id","served_model","language","item_id","type","subtype",
              "difficulty","gold","sample_idx","parsed_answer","correct","confidence",
              "parse_ok","finish_reason","reasoning_tokens","latency_s","raw_response","logprobs","error"]

# ---------- 解析 ----------
TRUE_WORDS = ("true","t","对","正确","yes","是")
FALSE_WORDS = ("false","f","错","错误","no","否")

def parse_response(text):
    """返回 (answer in {'True','False',None}, confidence in 1..5 or None, parse_ok bool)"""
    if not text:
        return None, None, False
    ans = conf = None
    ok = False
    # 去掉 markdown 代码围栏(小模型常把 JSON 包在 ```json ... ``` 里)
    t = re.sub(r"```[a-zA-Z]*", " ", text).replace("```", " ")
    # 1) 优先严格 JSON(完整闭合时)
    m = re.search(r"\{.*?\}", t, re.S)
    if m:
        try:
            d = json.loads(m.group(0))
            a = str(d.get("answer", "")).strip().lower()
            if any(a == w or a.startswith(w) for w in TRUE_WORDS): ans = "True"
            elif any(a == w or a.startswith(w) for w in FALSE_WORDS): ans = "False"
            c = d.get("confidence")
            if isinstance(c, bool):
                pass
            elif isinstance(c, (int, float)) and 1 <= int(c) <= 5:
                conf = int(c)
            elif isinstance(c, str) and c.strip()[:1] in "12345":
                conf = int(c.strip()[0])
        except Exception:
            pass
    # 2) 键值回退:即使 JSON 被截断(无闭合括号)也能抽出 answer/confidence
    if ans is None:
        ma = re.search(r'["\']?answer["\']?\s*[:=]\s*["\']?\s*'
                       r'(true|false|对|错|正确|错误|yes|no)', t, re.I)
        if ma:
            a = ma.group(1).lower()
            ans = "True" if any(a.startswith(w) for w in TRUE_WORDS) else "False"
    if conf is None:
        mc = re.search(r'["\']?confidence["\']?\s*[:=]\s*["\']?\s*([1-5])', t, re.I)
        if mc: conf = int(mc.group(1))
    # 3) 关键词兜底
    if ans is None:
        tl = t.lower()
        has_t = bool(re.search(r"\btrue\b|正确|(?<![不])对", tl))
        has_f = bool(re.search(r"\bfalse\b|错误|错", tl))
        if has_t and not has_f: ans = "True"
        elif has_f and not has_t: ans = "False"
    if conf is None:
        m2 = re.search(r"(?:confidence|确信|信心)\D{0,6}([1-5])", t, re.I)
        if m2: conf = int(m2.group(1))
    ok = ans is not None and conf is not None  # parse_ok = 是否成功拿到(答案,置信度)
    return ans, conf, ok

# ---------- 数据 ----------
def load_items(xlsx_path, sheet=None):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    items = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        items.append({
            "id": str(r[0]).strip(), "zh": r[1], "en": r[2],
            "gold": "True" if str(r[3]).strip().upper() == "T" else "False",
            "type": (str(r[4]).split()[0] if r[4] else ""),
            "subtype": r[5] or "", "difficulty": (str(r[6]).split()[0] if r[6] else ""),
        })
    return items

def _row_key(row):
    return (row["model"], row["language"], row["item_id"], row["sample_idx"])

def _row_valid(row):
    """一格是否算“有效完成”:成功解析且无调用错误。"""
    ok = str(row.get("parse_ok", "")).strip().lower() == "true"
    no_err = not str(row.get("error", "")).strip()
    return ok and no_err

def read_rows(csv_path):
    if not os.path.exists(csv_path):
        return []
    with open(csv_path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def load_done(csv_path, valid_only=False):
    """返回已完成的 (model,lang,item,sample) 键集合。
    valid_only=True 时,只有“有效完成”的格子才算数(失败/空回复会被重跑)。"""
    done = set()
    for row in read_rows(csv_path):
        if valid_only and not _row_valid(row):
            continue
        done.add(_row_key(row))
    return done

def expected_keys(name, languages, items, n):
    keys = set()
    for lang in languages:
        for it in items:
            if it[LANG_COL[lang]] is None:
                continue
            for k in range(n):
                keys.add((name, lang, it["id"], str(k)))
    return keys

def print_status(models, languages, items, cfg, done, extra=""):
    """打印每个模型的完成度,并返回总剩余数。"""
    print(f"\n===== 完成度{('('+extra+')') if extra else ''} =====")
    total_remaining = 0
    for mcfg in models:
        n = mcfg.get("n_samples", cfg.get("n_samples", 5))
        exp = expected_keys(mcfg["name"], languages, items, n)
        d = len(exp & done)
        rem = len(exp) - d
        total_remaining += rem
        flag = "✅ 完成" if rem == 0 else f"⬜ 剩 {rem}"
        print(f"  {mcfg['name']:<28} {d}/{len(exp)}  {flag}")
    if total_remaining == 0:
        print("  →  全部模型已测试完成 ✓")
    else:
        print(f"  →  仍有 {total_remaining} 条未完成;重跑同一命令会自动从断点继续。")
    return total_remaining

# ---------- 调用 ----------
def is_reasoning_model(model_id):
    """OpenAI 的 GPT-5 / o 系列推理模型:需 max_completion_tokens、不接受自定义 temperature、不支持 logprobs。"""
    m = (model_id or "").lower()
    return m.startswith(("gpt-5", "o1", "o3", "o4")) or "reason" in m

def _build_kwargs(mcfg, messages, seed, capture_logprobs, reasoning):
    kw = dict(model=mcfg["model_id"], messages=messages)
    tok = mcfg.get("max_tokens", 1024)
    if reasoning:
        kw["max_completion_tokens"] = tok           # 推理模型用这个字段
        # 不传 temperature(推理模型仅支持默认值);不传 logprobs(不支持)
    else:
        kw["temperature"] = mcfg.get("temperature", 0.7)
        kw["max_tokens"] = tok
        if capture_logprobs and mcfg.get("supports_logprobs"):
            kw["logprobs"] = True
            kw["top_logprobs"] = 5
    if seed is not None:
        kw["seed"] = seed
    if mcfg.get("json_mode"):
        kw["response_format"] = {"type": "json_object"}
    if mcfg.get("extra_body"):                       # 厂商专属参数透传(如 Qwen 的 enable_thinking)
        kw["extra_body"] = mcfg["extra_body"]
    return kw

def _once(client, kwargs, mcfg, capture_logprobs):
    """执行一次请求(流式或非流式),统一返回 (content, finish_reason, reasoning_tokens, served, logprobs)。"""
    if mcfg.get("stream"):                            # 部分思考型模型(如 Qwen)开思考需流式
        kwargs = dict(kwargs)
        kwargs["stream"] = True
        kwargs["stream_options"] = {"include_usage": True}
        parts = []; fr = ""; served = ""; rt = 0
        for chunk in client.chat.completions.create(**kwargs):
            if getattr(chunk, "model", None):
                served = chunk.model
            if chunk.choices:
                d = chunk.choices[0].delta
                if getattr(d, "content", None):
                    parts.append(d.content)           # 只累积最终答案,不累积 reasoning
                if chunk.choices[0].finish_reason:
                    fr = chunk.choices[0].finish_reason
            u = getattr(chunk, "usage", None)
            if u:
                try: rt = u.completion_tokens_details.reasoning_tokens or 0
                except Exception: pass
        return "".join(parts), fr, rt, served, None
    resp = client.chat.completions.create(**kwargs)
    ch = resp.choices[0]
    content = ch.message.content or ""
    fr = getattr(ch, "finish_reason", None) or ""
    rt = 0
    try: rt = resp.usage.completion_tokens_details.reasoning_tokens or 0
    except Exception: pass
    served = getattr(resp, "model", None) or ""
    lp = None
    try:
        if capture_logprobs and ch.logprobs:
            lp = json.dumps(ch.logprobs.model_dump(), ensure_ascii=False)
    except Exception:
        lp = None
    return content, fr, rt, served, lp

def call_model(client, mcfg, messages, capture_logprobs=False, seed=None):
    # 优先按配置 reasoning;否则按 model_id 自动判断
    reasoning = mcfg.get("reasoning", is_reasoning_model(mcfg["model_id"]))
    kwargs = _build_kwargs(mcfg, messages, seed, capture_logprobs, reasoning)
    try:
        content, fr, rt, served, lp = _once(client, kwargs, mcfg, capture_logprobs)
    except Exception as e:
        # 参数不兼容(temperature / max_tokens / max_completion_tokens / unsupported)时,
        # 翻转推理风格再试一次。400 报错不产生 token、不额外计费。
        msg = str(e).lower()
        if any(w in msg for w in ("temperature", "max_tokens", "max_completion_tokens",
                                  "unsupported", "logprobs")):
            kwargs = _build_kwargs(mcfg, messages, seed, capture_logprobs, not reasoning)
            content, fr, rt, served, lp = _once(client, kwargs, mcfg, capture_logprobs)
        else:
            raise
    text = content if content else f"[EMPTY finish_reason={fr} reasoning_tokens={rt}]"
    return text, lp, served, fr, rt

def make_client(mcfg):
    api_key = os.environ.get(mcfg.get("api_key_env", ""), mcfg.get("api_key", "not-needed"))
    return OpenAI(base_url=mcfg["base_url"], api_key=api_key,
                  timeout=mcfg.get("timeout", 120))

def worker(client, mcfg, cfg, cap, lang, it, k):
    """执行一次调用,返回可直接写入 CSV 的行 dict。线程安全:不触碰共享写入。"""
    stmt = it[LANG_COL[lang]]
    messages = [{"role":"system","content":SYS[lang]},
                {"role":"user","content":USER[lang].format(s=stmt)}]
    seed = (cfg.get("base_seed", 1000) + k) if cfg.get("use_seed") else None
    err = ""; text = ""; lp = None; served = ""; fr = ""; rt = 0; t0 = time.time()
    for attempt in range(cfg.get("retries", 3)):
        try:
            text, lp, served, fr, rt = call_model(client, mcfg, messages, cap, seed)
            err = ""; break
        except Exception as e:
            err = f"{type(e).__name__}: {e}"
            time.sleep(2 * (attempt + 1))
    lat = round(time.time() - t0, 2)
    ans, conf, ok = parse_response(text)
    correct = "" if ans is None else str(ans == it["gold"])
    s = mcfg.get("sleep", cfg.get("sleep", 0.0))
    if s: time.sleep(s)
    return {
        "timestamp": datetime.datetime.now().isoformat(timespec="seconds"),
        "model": mcfg["name"], "model_id": mcfg["model_id"],
        "served_model": served, "language": lang,
        "item_id": it["id"], "type": it["type"], "subtype": it["subtype"],
        "difficulty": it["difficulty"], "gold": it["gold"], "sample_idx": k,
        "parsed_answer": ans or "", "correct": correct,
        "confidence": conf if conf is not None else "",
        "parse_ok": ok, "finish_reason": fr, "reasoning_tokens": rt, "latency_s": lat,
        "raw_response": (text or "").replace("\n", " ").strip()[:1000],
        "logprobs": lp or "", "error": err,
    }

# ---------- 主流程 ----------
def run(cfg, args):
    items = load_items(cfg["item_bank"], cfg.get("sheet"))
    if args.items:
        want_ids = [x.strip() for x in args.items.split(",")]
        idx = {it["id"]: it for it in items}
        items = [idx[i] for i in want_ids if i in idx]
    if args.limit:
        items = items[:args.limit]
    languages = args.languages.split(",") if args.languages else cfg.get("languages", ["zh","en"])
    models = cfg["models"]
    if args.models:
        want = set(args.models.split(","))
        models = [m for m in models if m["name"] in want]
        if not models:
            sys.exit(f"未找到指定模型:{args.models}")

    out_csv = os.path.join(cfg.get("output_dir", str(HERE / "results")),
                           args.output_file or cfg.get("output_file", "results.csv"))
    os.makedirs(os.path.dirname(out_csv), exist_ok=True)
    # 若已存在的 CSV 列结构与当前不一致(如旧版无 served_model 列),拒绝追加以免错位
    if os.path.exists(out_csv):
        with open(out_csv, newline="", encoding="utf-8") as f:
            hdr = next(csv.reader(f), [])
        if hdr and hdr != CSV_FIELDS:
            sys.exit(f"已存在的 {out_csv} 列结构与当前版本不一致(可能是旧数据)。\n"
                     f"请删除或改名该文件,或在 models.yaml 里把 output_file 改成新名字后重跑。")
    # --retry-failed:把失败/空回复的格子清出去重跑(先备份,再只保留“有效完成”的行,避免重复)
    if args.retry_failed and os.path.exists(out_csv):
        allrows = read_rows(out_csv)
        good = [r for r in allrows if _row_valid(r)]
        dropped = len(allrows) - len(good)
        if dropped:
            import shutil
            shutil.copy(out_csv, out_csv + ".bak")
            with open(out_csv, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=CSV_FIELDS); w.writeheader(); w.writerows(good)
            print(f"[retry-failed] 清出 {dropped} 条失败/无效行待重跑(已备份到 {out_csv}.bak)")

    # 完成判定:默认按“有效完成”计;--no-resume 则全部重跑
    valid_only = not args.count_any
    done = set() if args.no_resume else load_done(out_csv, valid_only=valid_only)

    # 完成度报告(运行前)
    remaining = print_status(models, languages, items, cfg, done, extra="运行前")
    if args.status:   # 只报告、不运行
        return
    if remaining == 0 and not args.no_resume:
        print("\n无待跑任务:所有配置的模型 × 语言 × 题 × 采样均已完成。")
        return

    new_file = not os.path.exists(out_csv)
    fout = open(out_csv, "a", newline="", encoding="utf-8")
    writer = csv.DictWriter(fout, fieldnames=CSV_FIELDS)
    if new_file:
        writer.writeheader(); fout.flush()

    write_lock = threading.Lock()
    for mcfg in models:
        mcfg.setdefault("max_tokens", cfg.get("max_tokens", 96))
        mcfg.setdefault("json_mode", cfg.get("json_mode", False))
        client = make_client(mcfg)
        n = mcfg.get("n_samples", cfg.get("n_samples", 5))
        cap = cfg.get("capture_logprobs", False)
        conc = int(args.concurrency or mcfg.get("concurrency", cfg.get("concurrency", 1)))
        # 组装待办任务(跳过已完成的)
        tasks = []
        for lang in languages:
            for it in items:
                if it[LANG_COL[lang]] is None:
                    continue
                for k in range(n):
                    if (mcfg["name"], lang, it["id"], str(k)) in done:
                        continue
                    tasks.append((lang, it, k))
        print(f"\n=== 模型 {mcfg['name']} ({mcfg['model_id']}) | 每题 {n} 次采样 | 并发 {conc} | 待跑 {len(tasks)} ===")
        cnt = 0
        with ThreadPoolExecutor(max_workers=conc) as ex:
            futs = [ex.submit(worker, client, mcfg, cfg, cap, lang, it, k)
                    for (lang, it, k) in tasks]
            for fut in as_completed(futs):
                row = fut.result()
                with write_lock:
                    writer.writerow(row); fout.flush()
                cnt += 1
                if cnt % 20 == 0:
                    print(f"  ...{cnt}/{len(tasks)}", flush=True)
        print(f"  [{mcfg['name']}] 完成 {cnt} 次调用")
    fout.close()
    # 完成度报告(运行后):重新读盘统计,判断整批是否测完
    done_after = load_done(out_csv, valid_only=valid_only)
    print(f"\n结果:{out_csv}")
    print_status(models, languages, items, cfg, done_after, extra="运行后")

# ---------- 自测 ----------
def selftest():
    cases = [
        ('{"answer": "True", "confidence": 4}', ("True", 4, True)),
        ('{"answer":"False","confidence":1}', ("False", 1, True)),
        ('  {"answer": "对", "confidence": "3"} ', ("True", 3, True)),
        ('The answer is False. Confidence: 2', ("False", 2, True)),
        ('```json { "answer": "True", "confidence":', ("True", None, False)),
        ('```json\n{"answer":"False","confidence":3}\n```', ("False", 3, True)),
        ('True', ("True", None, False)),
        ('随便乱说没有答案', (None, None, False)),
    ]
    ok_all = True
    for text, exp in cases:
        got = parse_response(text)
        flag = "OK " if got == exp else "FAIL"
        if got != exp: ok_all = False
        print(f"[{flag}] {text[:40]!r} -> {got} (expect {exp})")
    print("解析器自测:", "全部通过" if ok_all else "存在失败,请检查")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", default=str(HERE / "models.yaml"))
    ap.add_argument("--models", help="逗号分隔的模型 name 过滤")
    ap.add_argument("--languages", help="逗号分隔,如 zh 或 en 或 zh,en")
    ap.add_argument("--limit", type=int, help="只跑前 N 题(自测用)")
    ap.add_argument("--items", help="只跑指定题号,逗号分隔,如 O01,D50,T08,H16(分层 pilot 用)")
    ap.add_argument("--concurrency", type=int, help="并发请求数(覆盖配置);本地模型建议 ≤ LM Studio 的 Max Concurrent Predictions")
    ap.add_argument("--output-file", dest="output_file", help="覆盖输出文件名(如 results_deepseek.csv),便于 API 各家分开存")
    ap.add_argument("--no-resume", action="store_true", help="不跳过已完成项(全部重跑)")
    ap.add_argument("--status", action="store_true", help="只报告完成度并退出,不调用模型")
    ap.add_argument("--retry-failed", dest="retry_failed", action="store_true",
                    help="把失败/空回复的格子清出重跑(先备份 .bak)")
    ap.add_argument("--count-any", dest="count_any", action="store_true",
                    help="只要有记录就算完成(默认只把“解析成功且无错误”的算完成)")
    ap.add_argument("--selftest", action="store_true", help="仅测试解析器,不调用模型")
    args = ap.parse_args()
    if args.selftest:
        selftest(); return
    load_dotenv(HERE / ".env")          # 自动加载本目录 .env 里的 API 密钥
    with open(args.config, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    run(cfg, args)

if __name__ == "__main__":
    main()
