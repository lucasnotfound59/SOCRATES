#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sample_items.py — 人类被试问卷:分层随机抽样器
从题库为每个被试抽一份题,保证【4 类型 × 3 难度 = 12 格】每格抽相同数量,
因此每个类别、每个难度都严格均匀。

要点:
- 每格抽 K 题(--per-cell K)→ 每份共 12K 题;默认 K=3 → 36 题(接近 40 且完美均匀)。
- 问卷【不含】题型/难度/答案(避免暗示被试);答案与分层信息另存 answer_key.csv。
- 每个被试独立随机抽,题序打乱;--seed 保证可复现。

用法:
  pip install openpyxl
  # 生成 8 份英文问卷,每格 3 题(=每份36题):
  python3 sample_items.py --n 8 --per-cell 3 --language en --seed 42
  # 中文、每份48题(每格4):
  python3 sample_items.py --n 8 --per-cell 4 --language zh
输出到 human_samples/:
  participant_01.xlsx ...(发给被试)
  answer_key.csv     (研究者评分/合并用)
  sampling_report.txt(抽样均衡性记录)
"""
import argparse, csv, os, random, datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
TYPES = ["常规", "学科", "陷阱", "幻觉"]
DIFFS = ["易", "中", "难"]
TYPE_EN = {"常规": "Ordinary", "学科": "Discipline", "陷阱": "Trap", "幻觉": "Hallucination"}

SCALE = {
    "en": ["1 = pure guess", "2 = not very sure", "3 = half-and-half",
           "4 = fairly sure", "5 = very sure"],
    "zh": ["1 = 完全靠猜", "2 = 不太确定", "3 = 一半一半",
           "4 = 比较确定", "5 = 非常确定"],
}
INSTR = {
    "en": ("For each statement, decide whether it is True or False, then rate how confident "
           "you are using the 1–5 scale below. Answer only from your own knowledge; you will "
           "not be told whether you are correct."),
    "zh": ("对每个陈述,请判断它是「对(True)」还是「错(False)」,然后用下面的 1–5 档报告你的确信程度。"
           "仅凭你自己的知识作答,你不会被告知对错。"),
}
COLS = {
    "en": ["#", "Item", "Statement", "Your answer (True/False)", "Confidence (1-5)"],
    "zh": ["#", "题号", "陈述", "你的判断(对/错)", "确信度(1-5)"],
}


def load_items(xlsx_path, sheet=None):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    items = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r or not r[0]:
            continue
        items.append({
            "id": str(r[0]).strip(), "zh": r[1], "en": r[2],
            "gold": "True" if str(r[3]).strip().upper() == "T" else "False",
            "type": str(r[4]).split()[0] if r[4] else "",
            "difficulty": str(r[6]).split()[0] if r[6] else "",
        })
    return items


def build_cells(items):
    cells = {(t, d): [] for t in TYPES for d in DIFFS}
    for it in items:
        key = (it["type"], it["difficulty"])
        if key in cells:
            cells[key].append(it)
    return cells


def sample_one(cells, per_cell, rng):
    """每格随机抽 per_cell 题,合并后打乱题序。"""
    chosen = []
    for key in cells:
        pool = cells[key]
        if len(pool) < per_cell:
            raise ValueError(f"格子 {key} 只有 {len(pool)} 题,不足 per_cell={per_cell}")
        chosen.extend(rng.sample(pool, per_cell))
    rng.shuffle(chosen)
    return chosen


def write_questionnaire(path, pid, chosen, language):
    wb = Workbook()
    ws = wb.active
    ws.title = "Questionnaire"
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = "Metacognition Questionnaire" if language == "en" else "元认知问卷"
    ws["A1"] = f"{title} — Participant {pid:02d}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A2"] = INSTR[language]
    ws["A2"].font = Font(name="Arial", size=10)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:E2"); ws.row_dimensions[2].height = 46
    ws["A3"] = ("Confidence scale:  " if language == "en" else "确信度五档:  ") + "   ".join(SCALE[language])
    ws["A3"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws.merge_cells("A3:E3")

    hdr_row = 5
    for j, c in enumerate(COLS[language], 1):
        cell = ws.cell(row=hdr_row, column=j, value=c)
        cell.fill = hfill; cell.font = hfont; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, it in enumerate(chosen, 1):
        stmt = it[language]
        row = hdr_row + i
        vals = [i, it["id"], stmt, "", ""]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(j == 3),
                                       horizontal="left" if j == 3 else "center")
    for col, w in zip("ABCDE", [5, 9, 66, 20, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{hdr_row+1}"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--bank", default=str(HERE / ".." / "题库_ItemBank_v2_400.xlsx"))
    ap.add_argument("--sheet", default="ItemBank 题库")
    ap.add_argument("--n", type=int, default=1, help="生成多少份(被试数)")
    ap.add_argument("--per-cell", dest="per_cell", type=int, default=3,
                    help="每个(类型×难度)格子抽几题;每份共 12×该值。默认3→36题")
    ap.add_argument("--language", choices=["en", "zh"], default="en")
    ap.add_argument("--seed", type=int, default=None, help="随机种子(可复现);留空则用系统随机")
    ap.add_argument("--out-dir", dest="out_dir", default=str(HERE / "human_samples"))
    args = ap.parse_args()

    items = load_items(args.bank, args.sheet)
    cells = build_cells(items)
    smallest = min(len(v) for v in cells.values())
    if args.per_cell > smallest:
        raise SystemExit(f"per_cell={args.per_cell} 超过最小格子容量 {smallest};请调小。")

    os.makedirs(args.out_dir, exist_ok=True)
    seed = args.seed if args.seed is not None else random.randrange(1 << 30)
    total_per = 12 * args.per_cell

    key_rows = []
    for p in range(1, args.n + 1):
        rng = random.Random(seed * 1_000_003 + p)   # 每份独立、可复现(int 种子,兼容 Python 3.14)
        chosen = sample_one(cells, args.per_cell, rng)
        qpath = os.path.join(args.out_dir, f"participant_{p:02d}.xlsx")
        write_questionnaire(qpath, p, chosen, args.language)
        for order, it in enumerate(chosen, 1):
            key_rows.append({
                "participant_id": p, "order": order, "item_id": it["id"],
                "type": it["type"], "difficulty": it["difficulty"],
                "gold": it["gold"], "language": args.language,
            })

    keypath = os.path.join(args.out_dir, "answer_key.csv")
    with open(keypath, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["participant_id", "order", "item_id",
                                          "type", "difficulty", "gold", "language"])
        w.writeheader(); w.writerows(key_rows)

    # 均衡性报告
    import collections
    report = [f"抽样报告 {datetime.datetime.now().isoformat(timespec='seconds')}",
              f"题库: {args.bank}", f"被试数: {args.n} | 每格 {args.per_cell} 题 | 每份 {total_per} 题",
              f"语言: {args.language} | seed: {seed}", ""]
    # 每份的格子分布(应全部等于 per_cell)
    per_cell_counts = collections.Counter((r["participant_id"], r["type"], r["difficulty"]) for r in key_rows)
    uniform = all(v == args.per_cell for v in per_cell_counts.values())
    report.append(f"每份每格数量是否全部 = {args.per_cell}: {'是 ✓ (完全均匀)' if uniform else '否 ✗'}")
    # 每份 T/F 平衡
    tf = collections.Counter((r["participant_id"], r["gold"]) for r in key_rows)
    report.append("每份 T/F(答案键未强制均衡,仅记录):")
    for p in range(1, args.n + 1):
        report.append(f"  被试{p:02d}: T={tf[(p,'True')]} F={tf[(p,'False')]}")
    rp = os.path.join(args.out_dir, "sampling_report.txt")
    open(rp, "w", encoding="utf-8").write("\n".join(report))

    print(f"生成 {args.n} 份问卷 → {args.out_dir}/participant_XX.xlsx")
    print(f"答案键 → {keypath}")
    print(f"每份 {total_per} 题,每(类型×难度)格 {args.per_cell} 题,{'完全均匀 ✓' if uniform else '不均匀 ✗'}")
    print(f"报告 → {rp}")


if __name__ == "__main__":
    main()
